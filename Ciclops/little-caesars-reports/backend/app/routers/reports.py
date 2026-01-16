"""
Reports Router - Little Caesars Reports
Elena + Julia: "Aquí generamos los reportes visuales y datos para el dashboard"
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, date
import io

from app.models import DashboardData, FinancialDataResponse
from app.services import (
    get_firebase_service, FirebaseService,
    get_claude_service, ClaudeService
)
from app.routers.auth import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/dashboard")
async def get_dashboard_data(
    current_user: dict = Depends(get_current_user),
    firebase: FirebaseService = Depends(get_firebase_service)
):
    """
    Obtener datos para el dashboard principal
    Elena: "Los KPIs bonitos que van arriba"
    """
    try:
        franchise_id = current_user.get("franchise_id", "default")
        current_period = datetime.utcnow().strftime("%Y-%m")

        # Obtener datos del periodo actual
        current_data = await firebase.query_documents(
            "financial_data",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("period", "==", current_period)
            ],
            limit=1
        )

        # Obtener datos del periodo anterior para comparar
        prev_month = datetime.utcnow().month - 1 or 12
        prev_year = datetime.utcnow().year if prev_month != 12 else datetime.utcnow().year - 1
        prev_period = f"{prev_year}-{prev_month:02d}"

        previous_data = await firebase.query_documents(
            "financial_data",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("period", "==", prev_period)
            ],
            limit=1
        )

        # Calcular métricas
        if current_data:
            data = current_data[0].get("data", {})
            ingresos = data.get("ingresos", {})
            gastos = data.get("gastos", {})
            costos = data.get("costos", {})

            total_revenue = sum([
                ingresos.get("ventas_mostrador", 0),
                ingresos.get("ventas_delivery", 0),
                ingresos.get("ventas_app", 0),
                ingresos.get("otros", 0)
            ])

            total_costs = sum(costos.values()) if isinstance(costos, dict) else 0
            total_expenses = sum([
                gastos.get("nomina", 0),
                gastos.get("renta", 0),
                gastos.get("luz", 0),
                gastos.get("agua", 0),
                gastos.get("gas", 0),
                gastos.get("marketing", 0),
                gastos.get("mantenimiento", 0),
                gastos.get("otros", 0)
            ])

            net_profit = total_revenue - total_costs - total_expenses
            net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0

            # Calcular tendencia vs periodo anterior
            if previous_data:
                prev_ingresos = previous_data[0].get("data", {}).get("ingresos", {})
                prev_revenue = sum([
                    prev_ingresos.get("ventas_mostrador", 0),
                    prev_ingresos.get("ventas_delivery", 0),
                    prev_ingresos.get("ventas_app", 0),
                    prev_ingresos.get("otros", 0)
                ])
                if prev_revenue > 0:
                    trend = ((total_revenue - prev_revenue) / prev_revenue) * 100
                    revenue_trend = f"+{trend:.1f}%" if trend >= 0 else f"{trend:.1f}%"
                else:
                    revenue_trend = "N/A"
            else:
                revenue_trend = "N/A"

            return {
                "current_period": current_period,
                "total_revenue": total_revenue,
                "net_margin": round(net_margin, 1),
                "net_profit": net_profit,
                "revenue_trend": revenue_trend,
                "alerts": current_data[0].get("alerts", []),
                "revenue_by_channel": {
                    "Mostrador": ingresos.get("ventas_mostrador", 0),
                    "Delivery": ingresos.get("ventas_delivery", 0),
                    "App": ingresos.get("ventas_app", 0),
                    "Otros": ingresos.get("otros", 0)
                },
                "expenses_breakdown": {
                    "Nómina": gastos.get("nomina", 0),
                    "Renta": gastos.get("renta", 0),
                    "Servicios": gastos.get("luz", 0) + gastos.get("agua", 0) + gastos.get("gas", 0),
                    "Marketing": gastos.get("marketing", 0),
                    "Otros": gastos.get("mantenimiento", 0) + gastos.get("otros", 0)
                }
            }
        else:
            return {
                "current_period": current_period,
                "total_revenue": 0,
                "net_margin": 0,
                "net_profit": 0,
                "revenue_trend": "N/A",
                "alerts": [],
                "revenue_by_channel": {},
                "expenses_breakdown": {},
                "message": "No hay datos para este periodo. ¡Sube tu primer documento!"
            }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@router.get("/pnl")
async def get_profit_and_loss(
    period: Optional[str] = Query(None, description="Periodo en formato YYYY-MM"),
    current_user: dict = Depends(get_current_user),
    firebase: FirebaseService = Depends(get_firebase_service)
):
    """
    Obtener Estado de Resultados (P&L)
    Julia: "El reporte que todo gerente quiere ver"
    """
    try:
        franchise_id = current_user.get("franchise_id", "default")
        target_period = period or datetime.utcnow().strftime("%Y-%m")

        # Obtener datos del periodo
        data = await firebase.query_documents(
            "financial_data",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("period", "==", target_period)
            ],
            limit=10  # Pueden haber múltiples documentos
        )

        if not data:
            return {
                "period": target_period,
                "message": "No hay datos para este periodo"
            }

        # Agregar datos de múltiples documentos
        aggregated = {
            "ingresos": {"ventas_mostrador": 0, "ventas_delivery": 0, "ventas_app": 0, "otros": 0},
            "costos": {"harina": 0, "queso": 0, "pepperoni": 0, "vegetales": 0, "carnes": 0, "bebidas": 0, "otros": 0},
            "gastos": {"nomina": 0, "renta": 0, "luz": 0, "agua": 0, "gas": 0, "marketing": 0, "mantenimiento": 0, "otros": 0},
            "impuestos": {"iva": 0, "isr": 0, "imss": 0}
        }

        for doc in data:
            doc_data = doc.get("data", {})
            for category in aggregated:
                if category in doc_data:
                    for key in aggregated[category]:
                        aggregated[category][key] += doc_data[category].get(key, 0)

        # Calcular totales
        total_ingresos = sum(aggregated["ingresos"].values())
        total_costos = sum(aggregated["costos"].values())
        total_gastos = sum(aggregated["gastos"].values())
        total_impuestos = sum(aggregated["impuestos"].values())

        utilidad_bruta = total_ingresos - total_costos
        utilidad_operativa = utilidad_bruta - total_gastos
        utilidad_neta = utilidad_operativa - total_impuestos

        return {
            "period": target_period,
            "ingresos": {
                "detalle": aggregated["ingresos"],
                "total": total_ingresos
            },
            "costos": {
                "detalle": aggregated["costos"],
                "total": total_costos
            },
            "utilidad_bruta": utilidad_bruta,
            "margen_bruto": round((utilidad_bruta / total_ingresos * 100), 1) if total_ingresos > 0 else 0,
            "gastos": {
                "detalle": aggregated["gastos"],
                "total": total_gastos
            },
            "utilidad_operativa": utilidad_operativa,
            "margen_operativo": round((utilidad_operativa / total_ingresos * 100), 1) if total_ingresos > 0 else 0,
            "impuestos": {
                "detalle": aggregated["impuestos"],
                "total": total_impuestos
            },
            "utilidad_neta": utilidad_neta,
            "margen_neto": round((utilidad_neta / total_ingresos * 100), 1) if total_ingresos > 0 else 0
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@router.get("/sales")
async def get_sales_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: dict = Depends(get_current_user),
    firebase: FirebaseService = Depends(get_firebase_service)
):
    """
    Obtener reporte de ventas
    """
    try:
        franchise_id = current_user.get("franchise_id", "default")

        # Por defecto, último mes
        if not end_date:
            end_date = datetime.utcnow().date()
        if not start_date:
            start_date = date(end_date.year, end_date.month, 1)

        # Obtener documentos del periodo
        documents = await firebase.query_documents(
            "documents",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("type", "==", "sales_report"),
                ("status", "==", "completed")
            ],
            order_by="uploaded_at",
            limit=100
        )

        sales_data = []
        for doc in documents:
            extracted = doc.get("extracted_data", {})
            if extracted and extracted.get("datos", {}).get("ingresos"):
                sales_data.append({
                    "date": doc.get("uploaded_at"),
                    "document_id": doc.get("id"),
                    "ingresos": extracted["datos"]["ingresos"]
                })

        return {
            "period": {
                "start": start_date.isoformat(),
                "end": end_date.isoformat()
            },
            "sales": sales_data,
            "total_documents": len(sales_data)
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


@router.post("/export")
async def export_report(
    report_type: str = Query(..., description="Tipo de reporte: pnl, sales, dashboard"),
    format: str = Query("pdf", description="Formato: pdf, excel"),
    period: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    firebase: FirebaseService = Depends(get_firebase_service)
):
    """
    Exportar reporte a PDF o Excel
    Elena: "Te lo dejo bonito para presentar"
    """
    try:
        # Obtener datos según tipo de reporte
        if report_type == "pnl":
            data = await get_profit_and_loss(period, current_user, firebase)
        elif report_type == "dashboard":
            data = await get_dashboard_data(current_user, firebase)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Tipo de reporte no soportado"
            )

        if format == "pdf":
            # Generar PDF
            pdf_buffer = await _generate_pdf_report(report_type, data)
            return StreamingResponse(
                io.BytesIO(pdf_buffer),
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f"attachment; filename=reporte_{report_type}_{period or 'actual'}.pdf"
                }
            )
        elif format == "excel":
            # Generar Excel
            excel_buffer = await _generate_excel_report(report_type, data)
            return StreamingResponse(
                io.BytesIO(excel_buffer),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=reporte_{report_type}_{period or 'actual'}.xlsx"
                }
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Formato no soportado. Usa: pdf, excel"
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )


async def _generate_pdf_report(report_type: str, data: dict) -> bytes:
    """Genera un reporte PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Estilo personalizado para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor("#F15A22"),  # Naranja Little Caesars
        spaceAfter=30
    )

    # Título
    title = f"Reporte {report_type.upper()}"
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Little Caesars Reports", styles['Heading2']))
    elements.append(Spacer(1, 20))

    # Contenido según tipo de reporte
    if report_type == "pnl" and "ingresos" in data:
        # Tabla de ingresos
        elements.append(Paragraph("Ingresos", styles['Heading3']))
        ingresos_data = [["Concepto", "Monto"]]
        for key, value in data["ingresos"]["detalle"].items():
            ingresos_data.append([key.replace("_", " ").title(), f"${value:,.2f}"])
        ingresos_data.append(["TOTAL", f"${data['ingresos']['total']:,.2f}"])

        table = Table(ingresos_data, colWidths=[300, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#F15A22")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FFF5F0")),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#E0E0E0"))
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

        # Resumen
        elements.append(Paragraph("Resumen", styles['Heading3']))
        resumen_data = [
            ["Utilidad Bruta", f"${data.get('utilidad_bruta', 0):,.2f}"],
            ["Margen Bruto", f"{data.get('margen_bruto', 0)}%"],
            ["Utilidad Neta", f"${data.get('utilidad_neta', 0):,.2f}"],
            ["Margen Neto", f"{data.get('margen_neto', 0)}%"],
        ]
        table = Table(resumen_data, colWidths=[300, 150])
        table.setStyle(TableStyle([
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#E0E0E0"))
        ]))
        elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


async def _generate_excel_report(report_type: str, data: dict) -> bytes:
    """Genera un reporte Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {report_type}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F15A22", end_color="F15A22", fill_type="solid")

    # Título
    ws['A1'] = f"Reporte {report_type.upper()} - Little Caesars"
    ws['A1'].font = Font(bold=True, size=16, color="F15A22")
    ws.merge_cells('A1:D1')

    if report_type == "pnl" and "ingresos" in data:
        # Headers
        ws['A3'] = "Concepto"
        ws['B3'] = "Monto"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['B3'].font = header_font
        ws['B3'].fill = header_fill

        row = 4
        # Ingresos
        ws[f'A{row}'] = "INGRESOS"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        for key, value in data["ingresos"]["detalle"].items():
            ws[f'A{row}'] = key.replace("_", " ").title()
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1

        ws[f'A{row}'] = "Total Ingresos"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = data["ingresos"]["total"]
        ws[f'B{row}'].number_format = '$#,##0.00'
        ws[f'B{row}'].font = Font(bold=True)
        row += 2

        # Resumen
        ws[f'A{row}'] = "RESUMEN"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        resumen = [
            ("Utilidad Bruta", data.get("utilidad_bruta", 0)),
            ("Margen Bruto %", data.get("margen_bruto", 0)),
            ("Utilidad Neta", data.get("utilidad_neta", 0)),
            ("Margen Neto %", data.get("margen_neto", 0)),
        ]

        for label, value in resumen:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if "%" not in label:
                ws[f'B{row}'].number_format = '$#,##0.00'
            row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@router.get("/insights")
async def get_insights(
    period: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
    firebase: FirebaseService = Depends(get_firebase_service),
    claude_service: ClaudeService = Depends(get_claude_service)
):
    """
    Obtener insights generados por IA
    Julia: "Mi análisis profundo de tus números"
    """
    try:
        franchise_id = current_user.get("franchise_id", "default")
        target_period = period or datetime.utcnow().strftime("%Y-%m")

        # Obtener datos actuales
        current_data = await firebase.query_documents(
            "financial_data",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("period", "==", target_period)
            ],
            limit=10
        )

        if not current_data:
            return {
                "message": "No hay datos suficientes para generar insights"
            }

        # Obtener datos del periodo anterior
        prev_month = int(target_period.split("-")[1]) - 1 or 12
        prev_year = int(target_period.split("-")[0]) if prev_month != 12 else int(target_period.split("-")[0]) - 1
        prev_period = f"{prev_year}-{prev_month:02d}"

        previous_data = await firebase.query_documents(
            "financial_data",
            filters=[
                ("franchise_id", "==", franchise_id),
                ("period", "==", prev_period)
            ],
            limit=10
        )

        # Generar insights con Claude
        insights = await claude_service.generate_insights(
            financial_data=current_data[0] if current_data else {},
            previous_period=previous_data[0] if previous_data else None
        )

        return {
            "period": target_period,
            "insights": insights
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(e)
        )
